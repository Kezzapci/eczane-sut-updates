"""Build a compact, searchable medicine index from the official SGK EK-4A workbook."""
from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return str(value).replace("\u00a0", " ").strip()


def code(value: Any) -> str:
    raw = text(value)
    if raw.endswith(".0") and raw[:-2].isdigit():
        return raw[:-2]
    return raw


def split_barcodes(value: Any) -> list[str]:
    raw = code(value)
    if not raw:
        return []
    return [item for item in re.split(r"[,;|/\\\s]+", raw) if item]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next(
        index for index, row in enumerate(rows)
        if any("Güncel Barkod" in text(cell) for cell in row)
    )
    headers = [text(cell) for cell in rows[header_index]]
    positions = {header: index for index, header in enumerate(headers) if header}

    def get(row: tuple[Any, ...], *names: str) -> str:
        for name in names:
            for header, index in positions.items():
                if name.lower() in header.lower():
                    return text(row[index]) if index < len(row) else ""
        return ""

    items: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        name = get(row, "İlaç Adı")
        current_barcode = code(get(row, "Güncel Barkod"))
        public_no = code(get(row, "Kamu No"))
        if not name and not current_barcode:
            continue
        old_barcodes = split_barcodes(get(row, "Eski Barkodlar"))
        all_barcodes = list(dict.fromkeys([current_barcode, *old_barcodes]))
        items.append(
            {
                "publicNo": public_no,
                "barcode": current_barcode,
                "oldBarcodes": old_barcodes,
                "barcodes": all_barcodes,
                "name": name,
                "equivalenceGroup": get(row, "Eşdeğer İlaç Grubu"),
                "therapeuticGroup": get(row, "Terapötik Referans Grubu"),
                "entryDate": get(row, "Listeye Giriş Tarihi"),
                "activeDate": get(row, "Aktiflenme Tarihi"),
                "passiveDate": get(row, "Pasiflenme Tarihi"),
                "discountStatus": get(row, "Uygulanan İndirim Oranlarına Esas Durumu"),
                "specialDiscount": get(row, "Özel İskonto"),
                "pharmacistDiscount": get(row, "Eczacı İskonto"),
                "searchText": " ".join([public_no, current_barcode, *old_barcodes, name, get(row, "Eşdeğer İlaç Grubu"), get(row, "Terapötik Referans Grubu")]).casefold(),
            }
        )

    payload = {
        "schemaVersion": 1,
        "source": "SGK EK-4/A Bedeli Ödenecek İlaçlar Listesi",
        "sourceUrl": "https://www.sgk.gov.tr/Download/DownloadFile?f=0ec1109c-a3fb-4723-867e-20567d7a67f5.xlsx&d=fa049c02-7d15-412e-8fb8-430c4f4f8694",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceFile": args.workbook.name,
        "count": len(items),
        "items": items,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({"count": len(items), "output": str(args.output)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
