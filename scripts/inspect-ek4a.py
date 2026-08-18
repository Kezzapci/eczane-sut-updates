from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/Downloads/0ec1109c-a3fb-4723-867e-20567d7a67f5.xlsx')
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = ws.iter_rows(values_only=True)
for i, row in zip(range(8), rows):
    print(i, [str(value).strip() if value is not None else '' for value in row])
